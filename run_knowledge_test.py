#!/usr/bin/env python3
"""Run the H4 external-knowledge questions on four Qwen model modes.

Questions and reference answers come only from data/knowledge_questions.csv.
The reference answers are copied into the output workbook but are never shown
to the model. Each question is sent as a fresh, independent chat request.

The recommended entry point is:

    CUDA_VISIBLE_DEVICES=0 python run_knowledge_test.py --all

The parent process launches one subprocess per model/mode so vLLM and GPU
memory are cleanly released between model loads.
"""

from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
from multiprocessing import freeze_support
from pathlib import Path

try:
    from prompts import run_baseline2_causal as common
except ModuleNotFoundError:
    # Backward compatibility with the earlier project layout.
    import run_baseline2_causal as common


MODE_SPECS = (
    ("Qwen/Qwen3.5-9B", "off", "Qwen3.5-9B nonthinking"),
    ("Qwen/Qwen3.5-9B", "on", "Qwen3.5-9B thinking"),
    ("Qwen/Qwen3.5-27B", "off", "Qwen3.5-27B nonthinking"),
    ("Qwen/Qwen3.5-27B", "on", "Qwen3.5-27B thinking"),
)

CSV_FIELDS = (
    "example_number",
    "knowledge_test_question_1",
    "knowledge_test_answer_1",
    "knowledge_test_question_2",
    "knowledge_test_answer_2",
)

OUTPUT_HEADERS = (
    "Example Number",
    "Knowledge Test Question 1",
    "Knowledge Test Answer 1",
    "Knowledge Test Question 2",
    "Knowledge Test Answer 2",
    "Model Response 1",
    "Model Response 2",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run H4 knowledge questions with vLLM and write a four-tab XLSM."
        )
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument(
        "--all",
        action="store_true",
        help="Run 9B/27B in both thinking and nonthinking modes.",
    )
    target.add_argument("--model", choices=list(common.REVISIONS))
    parser.add_argument(
        "--thinking",
        choices=["on", "off"],
        help="Required with --model; omitted with --all.",
    )
    parser.add_argument(
        "--input",
        default="data/knowledge_questions.csv",
        help="CSV containing example_number and the four knowledge columns.",
    )
    parser.add_argument(
        "--output",
        default="results/knowledge_test/knowledge_test_responses.xlsm",
        help="Four-tab macro-enabled workbook to create or resume.",
    )
    parser.add_argument(
        "--xlsm-template",
        default="data/counterfactual_causal_propagation_examples_19.xlsm",
        help=(
            "Existing macro-enabled workbook used only as the XLSM container. "
            "Its data are not read; the questions come from --input."
        ),
    )
    parser.add_argument(
        "--overwrite-mode",
        action="store_true",
        help="Clear and regenerate responses for the selected model/mode.",
    )
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--batch-size", type=int, default=0)
    parser.add_argument("--max-new-tokens", type=int, default=None)
    parser.add_argument("--max-model-len", type=int, default=None)
    parser.add_argument("--tensor-parallel-size", type=int, default=1)
    parser.add_argument("--gpu-memory-utilization", type=float, default=0.90)
    parser.add_argument("--max-num-seqs", type=int, default=None)
    parser.add_argument(
        "--sampler-backend",
        choices=["native", "flashinfer"],
        default="native",
    )
    parser.add_argument(
        "--gdn-prefill-backend",
        choices=["triton", "flashinfer", "auto"],
        default="triton",
    )
    args = parser.parse_args()

    if args.all and args.thinking is not None:
        parser.error("Do not pass --thinking with --all.")
    if not args.all and args.thinking is None:
        parser.error("--thinking is required when --model is used.")
    if args.batch_size < 0:
        parser.error("--batch-size must be 0 or a positive integer.")
    if args.max_new_tokens is not None and args.max_new_tokens < 1:
        parser.error("--max-new-tokens must be at least 1.")
    if args.tensor_parallel_size < 1:
        parser.error("--tensor-parallel-size must be at least 1.")
    if not 0 < args.gpu_memory_utilization <= 1:
        parser.error("--gpu-memory-utilization must be in (0, 1].")
    if args.max_num_seqs is not None and args.max_num_seqs < 1:
        parser.error("--max-num-seqs must be at least 1.")
    return args


def nonempty(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_questions(path: Path) -> list[dict[str, str | int]]:
    if not path.exists():
        raise FileNotFoundError(f"Could not find knowledge-question CSV: {path}")

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames != list(CSV_FIELDS):
            raise ValueError(
                f"CSV columns must be exactly {list(CSV_FIELDS)}; "
                f"found {reader.fieldnames}."
            )

        rows: list[dict[str, str | int]] = []
        seen: set[int] = set()
        for csv_line, raw in enumerate(reader, start=2):
            try:
                example_number = int(nonempty(raw["example_number"]))
            except ValueError as error:
                raise ValueError(
                    f"Invalid example_number on CSV line {csv_line}."
                ) from error
            if example_number < 1:
                raise ValueError(
                    f"example_number must be positive on CSV line {csv_line}."
                )
            if example_number in seen:
                raise ValueError(f"Duplicate example_number: {example_number}")
            seen.add(example_number)

            row: dict[str, str | int] = {"example_number": example_number}
            for field in CSV_FIELDS[1:]:
                row[field] = nonempty(raw[field])

            for question_number in (1, 2):
                question = nonempty(row[f"knowledge_test_question_{question_number}"])
                answer = nonempty(row[f"knowledge_test_answer_{question_number}"])
                if bool(question) != bool(answer):
                    raise ValueError(
                        f"Example {example_number}, question {question_number}: "
                        "question and reference answer must either both be filled "
                        "or both be blank."
                    )
            rows.append(row)

    rows.sort(key=lambda item: int(item["example_number"]))
    if not rows:
        raise ValueError("The knowledge-question CSV contains no rows.")
    return rows


def tab_name(model: str, thinking: str) -> str:
    for expected_model, expected_thinking, sheet_name in MODE_SPECS:
        if model == expected_model and thinking == expected_thinking:
            return sheet_name
    raise ValueError(f"Unsupported model/mode: {model}, thinking={thinking}")


def atomic_save_workbook(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.name}.tmp.xlsm")
    workbook.save(temporary_path)
    os.replace(temporary_path, output_path)


def initialize_or_sync_workbook(
    *,
    questions: list[dict[str, str | int]],
    output_path: Path,
    template_path: Path,
):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required. Install it in ~/conda/llms with: "
            "python -m pip install openpyxl"
        ) from error

    if output_path.exists():
        workbook = load_workbook(output_path, keep_vba=True)
    else:
        if not template_path.exists():
            raise FileNotFoundError(
                "A real macro-enabled workbook is needed to produce a valid "
                f".xlsm file. Could not find template: {template_path}"
            )
        workbook = load_workbook(template_path, keep_vba=True)
        if workbook.vba_archive is None:
            raise ValueError(
                f"The template is named .xlsm but contains no VBA project: "
                f"{template_path}. Use a genuine macro-enabled workbook or "
                "change the requested output format to .xlsx."
            )
        for sheet in list(workbook.worksheets):
            workbook.remove(sheet)
        for _, _, sheet_name in MODE_SPECS:
            workbook.create_sheet(sheet_name)

    expected_names = [sheet_name for _, _, sheet_name in MODE_SPECS]
    for sheet_name in expected_names:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
    for sheet in list(workbook.worksheets):
        if sheet.title not in expected_names:
            workbook.remove(sheet)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    widths = (15, 56, 48, 56, 48, 56, 56)

    for _, _, sheet_name in MODE_SPECS:
        sheet = workbook[sheet_name]
        existing_responses: dict[int, tuple[str, str]] = {}
        if sheet.max_row >= 2:
            for row_number in range(2, sheet.max_row + 1):
                value = sheet.cell(row_number, 1).value
                try:
                    example_number = int(value)
                except (TypeError, ValueError):
                    continue
                existing_responses[example_number] = (
                    nonempty(sheet.cell(row_number, 6).value),
                    nonempty(sheet.cell(row_number, 7).value),
                )

        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
        sheet.append(OUTPUT_HEADERS)
        for item in questions:
            example_number = int(item["example_number"])
            response_1, response_2 = existing_responses.get(
                example_number, ("", "")
            )
            sheet.append(
                (
                    example_number,
                    item["knowledge_test_question_1"],
                    item["knowledge_test_answer_1"],
                    item["knowledge_test_question_2"],
                    item["knowledge_test_answer_2"],
                    response_1,
                    response_2,
                )
            )

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        sheet.row_dimensions[1].height = 34
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:G{len(questions) + 1}"
        sheet.sheet_view.showGridLines = False
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(1, column_number).column_letter
            ].width = width
        for row_number in range(2, len(questions) + 2):
            sheet.row_dimensions[row_number].height = 64
            for cell in sheet[row_number]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for desired_index, name in enumerate(expected_names):
        sheet = workbook[name]
        current_index = workbook.index(sheet)
        workbook.move_sheet(sheet, offset=desired_index - current_index)
    atomic_save_workbook(workbook, output_path)
    return workbook


def build_prompt(question: str) -> str:
    return f"""Answer the following factual question directly and concisely.
Do not mention this instruction or discuss the benchmark.

Question: {question}

Answer:"""


def run_all_modes(args: argparse.Namespace) -> int:
    base_command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "--input",
        args.input,
        "--output",
        args.output,
        "--xlsm-template",
        args.xlsm_template,
        "--seed",
        str(args.seed),
        "--batch-size",
        str(args.batch_size),
        "--tensor-parallel-size",
        str(args.tensor_parallel_size),
        "--gpu-memory-utilization",
        str(args.gpu_memory_utilization),
        "--sampler-backend",
        args.sampler_backend,
        "--gdn-prefill-backend",
        args.gdn_prefill_backend,
    ]
    if args.max_new_tokens is not None:
        base_command.extend(["--max-new-tokens", str(args.max_new_tokens)])
    if args.max_model_len is not None:
        base_command.extend(["--max-model-len", str(args.max_model_len)])
    if args.max_num_seqs is not None:
        base_command.extend(["--max-num-seqs", str(args.max_num_seqs)])
    if args.overwrite_mode:
        base_command.append("--overwrite-mode")

    for model, thinking, sheet_name in MODE_SPECS:
        print("\n" + "=" * 78, flush=True)
        print(f"Running {sheet_name}", flush=True)
        command = base_command + ["--model", model, "--thinking", thinking]
        subprocess.run(command, check=True)

    print("\nFinished all four model modes.")
    print("Workbook:", args.output)
    return 0


def run_one_mode(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    output_path = Path(args.output)
    template_path = Path(args.xlsm_template)
    questions = load_questions(input_path)
    workbook = initialize_or_sync_workbook(
        questions=questions,
        output_path=output_path,
        template_path=template_path,
    )

    sheet_name = tab_name(args.model, args.thinking)
    sheet = workbook[sheet_name]
    if args.overwrite_mode:
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, 6).value = None
            sheet.cell(row_number, 7).value = None
        atomic_save_workbook(workbook, output_path)

    requests: list[dict[str, object]] = []
    for row_number in range(2, sheet.max_row + 1):
        example_number = int(sheet.cell(row_number, 1).value)
        for question_number, question_column, response_column in (
            (1, 2, 6),
            (2, 4, 7),
        ):
            question = nonempty(sheet.cell(row_number, question_column).value)
            response = nonempty(sheet.cell(row_number, response_column).value)
            if question and not response:
                requests.append(
                    {
                        "example_number": example_number,
                        "question_number": question_number,
                        "row_number": row_number,
                        "response_column": response_column,
                        "question": question,
                    }
                )

    if not requests:
        print(f"No unfinished questions remain for {sheet_name}.")
        print("Workbook:", output_path)
        return 0

    thinking_enabled = args.thinking == "on"
    max_new_tokens = args.max_new_tokens
    if max_new_tokens is None:
        max_new_tokens = 4096 if thinking_enabled else 256
    max_model_len = args.max_model_len or (max_new_tokens + 2048)
    if max_model_len <= max_new_tokens:
        raise ValueError("--max-model-len must exceed --max-new-tokens.")
    max_num_seqs = args.max_num_seqs
    if max_num_seqs is None:
        max_num_seqs = 16 if args.model.endswith("9B") else 6

    print("Model:", args.model)
    print("Revision:", common.REVISIONS[args.model])
    print("Mode:", "thinking" if thinking_enabled else "nonthinking")
    print("Questions to run:", len(requests))
    print("Maximum new tokens:", max_new_tokens)
    print("Workbook tab:", sheet_name)
    print("Loading vLLM model...")

    os.environ.setdefault("HF_HUB_OFFLINE", "1")
    os.environ["VLLM_USE_FLASHINFER_SAMPLER"] = (
        "1" if args.sampler_backend == "flashinfer" else "0"
    )
    LLM, SamplingParams, _ = common.load_vllm()
    revision = common.REVISIONS[args.model]
    llm = LLM(
        model=args.model,
        revision=revision,
        tokenizer_revision=revision,
        tensor_parallel_size=args.tensor_parallel_size,
        dtype="bfloat16",
        max_model_len=max_model_len,
        gpu_memory_utilization=args.gpu_memory_utilization,
        max_num_seqs=max_num_seqs,
        seed=args.seed,
        language_model_only=True,
        additional_config={"gdn_prefill_backend": args.gdn_prefill_backend},
    )

    submission_size = args.batch_size or len(requests)
    for batch_start in range(0, len(requests), submission_size):
        batch = requests[batch_start : batch_start + submission_size]
        conversations = [
            [{"role": "user", "content": build_prompt(str(item["question"]))}]
            for item in batch
        ]
        seeds = [
            args.seed
            + int(item["example_number"]) * 10
            + int(item["question_number"])
            if thinking_enabled
            else None
            for item in batch
        ]
        sampling_params = [
            common.make_sampling_params(
                SamplingParams,
                thinking_enabled,
                max_new_tokens,
                request_seed,
            )
            for request_seed in seeds
        ]

        print(
            f"Submitting questions {batch_start + 1}-"
            f"{batch_start + len(batch)} of {len(requests)}..."
        )
        outputs = llm.chat(
            conversations,
            sampling_params=sampling_params,
            use_tqdm=True,
            chat_template_kwargs={"enable_thinking": thinking_enabled},
        )
        if len(outputs) != len(batch):
            raise RuntimeError(
                f"vLLM returned {len(outputs)} outputs for {len(batch)} questions."
            )

        for item, request_output in zip(batch, outputs):
            completion = request_output.outputs[0]
            raw_completion = completion.text.strip()
            token_count = len(completion.token_ids)
            exhausted = (
                completion.finish_reason == "length"
                or token_count >= max_new_tokens
            )
            _, response, thinking_status = common.split_qwen_completion(
                raw_completion,
                thinking_enabled,
                exhausted,
            )
            if response is None:
                response = (
                    "[NO FINAL ANSWER — "
                    + ("budget exhausted" if exhausted else thinking_status)
                    + "]"
                )
            sheet.cell(
                int(item["row_number"]),
                int(item["response_column"]),
            ).value = response
            print(
                f"Example {item['example_number']} Q{item['question_number']}: "
                f"{response}"
            )

        atomic_save_workbook(workbook, output_path)
        print("Saved progress to:", output_path)

    print(f"Finished {sheet_name}.")
    print("Workbook:", output_path)
    return 0


def main() -> int:
    args = parse_args()
    if args.all:
        return run_all_modes(args)
    return run_one_mode(args)


if __name__ == "__main__":
    freeze_support()
    raise SystemExit(main())
